from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login
from django.views.generic import View
from django.http import HttpRequest, JsonResponse, HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.db.models import Q, F, ExpressionWrapper, DecimalField, IntegerField, Case, When, Value, Count
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger, InvalidPage
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.core.validators import RegexValidator
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.conf import settings

from system_manage.decorators import permission_required
from system_manage.views.system_manage_views.auth_views import validate_birth, validate_phone
from system_manage.models import Customer, Order
from system_manage.services.customer_cache import get_cached_customers, get_cached_order_status_count

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from urllib.parse import quote



class OrderManageView(View):
    '''
        주문 관리 화면
    '''
    @method_decorator(permission_required(redirect_url='system_manage:denied'))
    def get(self, request: HttpRequest, *args, **kwargs):
        context = {}
        context['active_menu1'] = 'order'
        search_keyword = request.GET.get('search_keyword', '').strip()
        context['search_keyword'] = search_keyword

        paginate_by = '20'
        page = request.GET.get('page', '1')

        order = request.GET.get('order', 'desc')
        sort = request.GET.get('sort', 'created_at')
        status = request.GET.get('status', '0')

        excel = request.GET.get('excel')

        context['order'] = order
        context['sort'] = sort
        context['status'] = status

        context['customers'] = get_cached_customers()
        context['status_count'] = get_cached_order_status_count()

        if order == 'desc':
            ordering = [f'-{sort}', '-id']
        else:
            ordering = [f'{sort}', 'id']

        if status:
            query = Q(status=status)
        else:
            query = Q()

        if search_keyword:
            search_q = Q(customer__name__icontains=search_keyword)
            if search_keyword.isdigit():

                search_q |= Q(customer__phone=search_keyword)
            query &= search_q

        queryset = Order.objects.select_related('customer').filter(query).order_by(*ordering)
        if excel:
            return export_orders_excel(queryset)

        paginator = Paginator(queryset, paginate_by)
        try:
            page_obj = paginator.page(page)
        except (PageNotAnInteger, EmptyPage, InvalidPage):
            page = 1
            page_obj = paginator.page(page)

        pagelist = paginator.get_elided_page_range(page, on_each_side=3, on_ends=1)
        context['total_count'] = paginator.count
        context['pagelist'] = pagelist
        context['page_obj'] = page_obj
        context['last_page_number'] = paginator.num_pages        

        return render(request, 'customer_manage/order_manage.html', context)
    
    @method_decorator(permission_required(raise_exception=True))
    def post(self, request: HttpRequest, *args, **kwargs):
        order_name = request.POST['order_name'].strip()
        if order_name == '':
            return JsonResponse({'message': '주문명을 입력해주세요.'}, status=400)
        order_date = request.POST['order_date']
        if not validate_birth(order_date):
            return JsonResponse({'message': '주문날짜 형식 오류'}, status=400)
        order_note = request.POST['order_note'].strip()
        option = request.POST['option']
        if option not in ['0', '1', '2']:
            return JsonResponse({'message': '옵션 형식 오류'}, status=400)
        status = request.POST['status']
        if status not in ['0', '1', '2', '3']:
            return JsonResponse({'message': '상태 형식 오류'}, status=400)
        
        customer_name = request.POST['customer_name'].strip()
        customer_phone = request.POST['customer_phone'].strip()
        comment = request.POST.get('comment', '').strip()   
        
        try:
            with transaction.atomic():
                if customer_phone:
                    if not validate_phone(customer_phone):
                        return JsonResponse({'message': '전화번호 형식 오류'}, status=400)
                    customer, created = Customer.objects.get_or_create(phone=customer_phone)
                    if created:
                        customer.name = customer_name if customer_name else '이름없음'
                    elif customer_name and customer.name != customer_name:
                        customer.name = customer_name
                    customer.save()
                else:
                    customer = None

                Order.objects.create(
                    order_name = order_name,
                    order_date = order_date,
                    order_note = order_note,
                    option = option,
                    status = status,
                    customer = customer,
                    comment = comment
                )
        except:
            return JsonResponse({'message': '등록 오류'}, status=400)
        return JsonResponse({'message' : '등록 되었습니다.', 'url': reverse('system_manage:order_manage')},  status = 201)
    
@require_http_methods(["POST"])    
@permission_required(raise_exception=True)
def edit_order(request):
    order_id = int(request.POST['order_id'])
    try:
        order = Order.objects.get(id=order_id)
    except Order.DoesNotExist:
        return JsonResponse({'message': 'Order not found.'}, status=400)
    order_name = request.POST['order_name'].strip()
    if order_name == '':
        return JsonResponse({'message': '주문명을 입력해주세요.'}, status=400)
    order_date = request.POST['order_date']
    if not validate_birth(order_date):
        return JsonResponse({'message': '주문날짜 형식 오류'}, status=400)
    order_note = request.POST['order_note'].strip()
    option = request.POST['option']
    if option not in ['0', '1', '2']:
        return JsonResponse({'message': '옵션 형식 오류'}, status=400)
    status = request.POST['status']
    
    if status not in ['0', '1', '2', '3']:
        return JsonResponse({'message': '상태 형식 오류'}, status=400)
    
    customer_name = request.POST['customer_name'].strip()
    customer_phone = request.POST['customer_phone'].strip()
    comment = request.POST.get('comment', '').strip()   
    try:
        with transaction.atomic():
            if customer_phone:
                if not validate_phone(customer_phone):
                    return JsonResponse({'message': '전화번호 형식 오류'}, status=400)
                customer, created = Customer.objects.get_or_create(phone=customer_phone)
                if created:
                    customer.name = customer_name if customer_name else '이름없음'
                elif customer_name and customer.name != customer_name:
                    customer.name = customer_name
                customer.save()
            else:
                customer = None

            order.order_name = order_name
            order.order_date = order_date
            order.order_note = order_note
            order.option = option
            order.status =status
            order.comment = comment
            order.customer = customer
            order.save()
    except:
        return JsonResponse({'message': 'Error occurred while updating order.'}, status=400)

    return JsonResponse({'message': '수정 되었습니다.'}, status=200)

@require_http_methods(["POST"])    
@permission_required(raise_exception=True)
def delete_order(request):
    order_id = int(request.POST['order_id'])
    try:
        order = Order.objects.get(id=order_id)
    except Order.DoesNotExist:
        return JsonResponse({'message': 'Order not found.'}, status=400)
    try:
        order.delete()
    except:
        return JsonResponse({'message': 'Error occurred while deleting order.'}, status=400)

    return JsonResponse({'message': '삭제 되었습니다.'}, status=200)


@require_http_methods(["POST"])    
@permission_required(raise_exception=True)
def order_status(request):
    order_id = int(request.POST['order_id'])
    try:
        order = Order.objects.get(id=order_id)
    except Order.DoesNotExist:
        return JsonResponse({'message': 'Order not found.'}, status=400)
    status = request.POST['status']
    if status not in ['0', '1', '2', '3']:
        return JsonResponse({'message': '상태 형식 오류'}, status=400)
    try:
        with transaction.atomic():
            order.status =status
            order.save()
    except:
        return JsonResponse({'message': 'Error occurred while updating order.'}, status=400)
    status_count_dict = get_cached_order_status_count()

    return JsonResponse({'message': '업데이트 되었습니다.', 'status_count': status_count_dict}, status=200)



class CustomerOrderManageView(View):
    '''
        주문 관리 화면
    '''
    @method_decorator(permission_required(redirect_url='system_manage:denied'))
    def get(self, request: HttpRequest, *args, **kwargs):
        context = {}
        customer_id = kwargs.get('customer_id')
        customer = get_object_or_404(Customer, pk=customer_id)
        context['customer'] = customer
        
        context['active_menu1'] = 'order'
        search_keyword = request.GET.get('search_keyword', '').strip()
        context['search_keyword'] = search_keyword

        paginate_by = '20'
        page = request.GET.get('page', '1')

        order = request.GET.get('order', 'desc')
        sort = request.GET.get('sort', 'created_at')
        status = request.GET.get('status', '')

        excel = request.GET.get('excel')

        context['order'] = order
        context['sort'] = sort
        context['status'] = status

        if order == 'desc':
            ordering = [f'-{sort}', '-id']
        else:
            ordering = [f'{sort}', 'id']

        context['customers'] = get_cached_customers()


        if status:
            query = Q(customer=customer, status=status)
        else:
            query = Q(customer=customer)

        queryset = Order.objects.select_related('customer').filter(query).order_by(*ordering)
        if excel:
            return export_orders_excel(queryset)

        paginator = Paginator(queryset, paginate_by)
        try:
            page_obj = paginator.page(page)
        except (PageNotAnInteger, EmptyPage, InvalidPage):
            page = 1
            page_obj = paginator.page(page)

        pagelist = paginator.get_elided_page_range(page, on_each_side=3, on_ends=1)
        context['total_count'] = paginator.count
        context['pagelist'] = pagelist
        context['page_obj'] = page_obj
        context['last_page_number'] = paginator.num_pages        

        return render(request, 'customer_manage/customer_order_manage.html', context)


def export_orders_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = '주문목록'

    headers = [
        '주문날짜',
        '회원명',
        '전화번호',
        '주문내용',
        '중요',
        '잠금장치',
        '비고',
        '상태'
    ]
    ws.append(headers)

    # 헤더 스타일 + 컬럼 폭
    for col in range(1, len(headers) + 1):
        cell = ws[f'{get_column_letter(col)}1']
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = 20

    STATUS_MAP = {
        '0': '주문요청',
        '1': '주문완료',
        '2': '전달완료',
        '3': '주문취소',
    }

    OPTION_MAP = {
        '0': '없음',
        '1': '환산',
        '2': '별도',
    }
    for obj in queryset:
        ws.append([
            obj.order_date,
            obj.customer.name if obj.customer else '미등록',
            obj.customer.phone if obj.customer else '',
            obj.order_name,
            obj.order_note or '',
            OPTION_MAP.get(obj.option, '알수없음'),
            obj.comment or '',
            STATUS_MAP.get(obj.status, '알수없음'),
        ])

    filename = f'{settings.SITE_NAME}_주문내역_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    encoded_filename = quote(filename)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        "attachment; filename*=UTF-8''{}".format(encoded_filename)
    )

    wb.save(response)
    return response
